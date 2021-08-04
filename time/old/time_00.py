import shutil
import datetime
import os
from openpyxl import load_workbook

# determine filename based on today's date
today = datetime.date.today()
last_monday = today - datetime.timedelta(days=today.weekday())
file_name = f'Ward, Gratten 2021_Timesheet_{last_monday}.xlsm'
path = os.getcwd()
files = os.listdir(path)

def new_week(file_name):
    if file_name not in files:
        shutil.copy("Ward, Gratten 2021_Timesheet.xlsm", file_name)
        feedback = '\nNew file created.\n'
    else:
        feedback = "\nFile already exists.\n"
    return feedback

def add_task(file_name):
    wb = load_workbook(filename=file_name, read_only=False, keep_vba=True)
    ws = wb.worksheets[0]

    # find next row to populate
    cell_range = ws['A21':'A36']
    for cell in cell_range:
        if cell[0].value is None:
            empty = cell[0].row
            break
    else:
        print('Document full.')

    # collect user input
    project = input("Enter project: ")
    description = input("Enter description: ")
    seq = input("Enter sequence: ")
    act_code = input("Enter activity code: ")
    hours = input("Enter hours: ")

    # determine which day to populate hours

    # populate data
    ws.cell(row=empty, column=1).value = project
    ws.cell(row=empty, column=2).value = description
    ws.cell(row=empty, column=3).value = seq
    ws.cell(row=empty, column=4).value = act_code
    ws.cell(row=empty, column=5).value = hours
    wb.save(filename=TEST.xlsm)

    # notify user
    feedback = '\nTask recorded.\n'
    return feedback

selection = ''
while selection != 'E':
    selection = input('W - new week \n'
                      'T - add task\n'
                      'E - exit\n\n'
                      'Make a selection...')

    if selection == 'W':
        print(new_week(file_name))
    elif selection == 'T':
        print(add_task(file_name))